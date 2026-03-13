"""
LlamaIndex-powered document parser.

Replaces the bare PyMuPDF/openpyxl ad-hoc extraction with LlamaIndex's
file readers so we get clean, structured text for every supported format:
  • PDF   – PDFReader  (wraps PyMuPDF / pdfminer, page-level text)
  • Excel – PandasExcelReader  (reads every sheet, preserves numbers)
  • CSV   – PandasCSVReader
  • DOCX  – DocxReader
  • TXT/JSON – plain read

The single public function `parse_document(path, file_type)` returns a
plain-text string that is stored in `Document.parsed_text`.
"""

from __future__ import annotations

import io
import logging
import os
import re
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

_MAX_DOC_CHARS = int(os.environ.get("AIBAA_MAX_PARSED_CHARS", "1200000"))
_MAX_EXCEL_ROWS_PER_SHEET = int(os.environ.get("AIBAA_MAX_EXCEL_ROWS_PER_SHEET", "4000"))
_OCR_MIN_PAGE_CHARS = int(os.environ.get("AIBAA_OCR_MIN_PAGE_CHARS", "40"))

_NUMERIC_LINE_RE = re.compile(r"(?:\d[\d,\.\-]*%?)")
_FY_LABEL_RE = re.compile(r"^FY(\d{2})$", re.I)
_SECTION_PATTERNS: dict[str, tuple[re.Pattern[str], ...]] = {
    "Revenue": (
        re.compile(r"\brevenue\b", re.I),
        re.compile(r"revenue\s+from\s+operations", re.I),
        re.compile(r"\bnet\s+sales\b", re.I),
        re.compile(r"\btotal\s+income\b", re.I),
    ),
    "EBITDA": (
        re.compile(r"\bebitda\b", re.I),
        re.compile(r"operating\s+profit", re.I),
        re.compile(r"profit\s+before\s+(?:interest|finance)\s+and\s+(?:tax|depreciation)", re.I),
    ),
    "Debt": (
        re.compile(r"\btotal\s+debt\b", re.I),
        re.compile(r"\bborrowings?\b", re.I),
        re.compile(r"lease\s+liabilit", re.I),
        re.compile(r"long[-\s]?term\s+debt", re.I),
        re.compile(r"short[-\s]?term\s+borrowings?", re.I),
    ),
    "Cash": (
        re.compile(r"cash\s+and\s+cash\s+equivalents", re.I),
        re.compile(r"\bcash\s+equivalents\b", re.I),
        re.compile(r"bank\s+balances", re.I),
        re.compile(r"current\s+investments", re.I),
        re.compile(r"liquid\s+investments", re.I),
    ),
}

_SPREADSHEET_REVENUE_LABELS = (
    "total revenue",
    "revenue from operations",
    "net sales",
)
_SPREADSHEET_OPERATING_INCOME_LABELS = (
    "operating income",
    "operating profit",
)
_SPREADSHEET_DEPRECIATION_LABELS = (
    "depreciation",
    "depreciation and amortisation",
    "depreciation and amortization",
)
_SPREADSHEET_CASH_LABELS = (
    "cash and cash equivalents",
    "investments",
    "bank deposits",
)
_SPREADSHEET_DEBT_LABELS = (
    "short term borrowings",
    "short term debt borrowings",
    "short term debt",
    "long term debt borrowings",
    "long term borrowings",
    "borrowings",
)
_SPREADSHEET_PAT_LABELS = (
    "net profit after taxes",
    "profit after tax",
    "net income after extraordinary items",
    "net income before extraordinary items",
)
_SPREADSHEET_EPS_LABELS = (
    "basic and diluted eps",
    "basic eps",
)
_SPREADSHEET_SHARES_LABELS = (
    "weighted average no of shares used in computing basic and diluted eps",
    "weighted average number of shares used in computing basic and diluted eps",
    "weighted average shares",
)


def _normalize_label(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _to_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _match_any_label(label: str, candidates: tuple[str, ...]) -> bool:
    return any(candidate in label for candidate in candidates)


def _detect_reporting_unit(text_snippets: list[str]) -> tuple[str, float]:
    blob = " ".join(text_snippets).lower()
    if "crore" in blob or "crores" in blob:
        return "crores", 10_000_000.0
    if "lakh" in blob or "lakhs" in blob:
        return "lakhs", 100_000.0
    if "inr mn" in blob or "₹ mn" in blob or "rs mn" in blob or "million" in blob or " mn" in blob:
        return "millions", 1_000_000.0
    if "thousand" in blob or "000" in blob:
        return "thousands", 1_000.0
    return "absolute", 1.0


def _sheet_priority(sheet_name: str) -> int:
    name = sheet_name.lower()
    score = 0
    if "ifrs" in name:
        score += 40
    if "inr" in name:
        score += 80
    if "pnl" in name:
        score += 20
    if "bs" in name:
        score += 10
    if "usd" in name:
        score -= 100
    return score


def _find_fy_columns(ws) -> list[tuple[int, int]]:
    best_matches: list[tuple[int, int]] = []
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        current: list[tuple[int, int]] = []
        for idx, value in enumerate(row, start=1):
            match = _FY_LABEL_RE.fullmatch(str(value or "").strip())
            if not match:
                continue
            year = 2000 + int(match.group(1))
            current.append((idx, year))
        if len(current) > len(best_matches):
            best_matches = current
    return sorted(best_matches, key=lambda item: item[1])[-5:]


def _sum_series(rows: list[tuple[int, tuple[Any, ...]]], fy_cols: list[tuple[int, int]], multiplier: float) -> list[float]:
    totals = [0.0 for _ in fy_cols]
    for _, row in rows:
        for idx, (col_idx, _) in enumerate(fy_cols):
            value = _to_number(row[col_idx - 1] if col_idx - 1 < len(row) else None)
            if value is not None:
                totals[idx] += value * multiplier
    return totals


def _series_from_row(row: tuple[Any, ...], fy_cols: list[tuple[int, int]], multiplier: float = 1.0) -> list[Optional[float]]:
    values: list[Optional[float]] = []
    for col_idx, _ in fy_cols:
        raw = row[col_idx - 1] if col_idx - 1 < len(row) else None
        value = _to_number(raw)
        values.append((value * multiplier) if value is not None else None)
    return values


def _pick_row(
    row_entries: list[tuple[int, str, tuple[Any, ...]]],
    candidates: tuple[str, ...],
) -> Optional[tuple[int, str, tuple[Any, ...]]]:
    exact_match = next(
        (entry for entry in row_entries if entry[1] in candidates),
        None,
    )
    if exact_match:
        return exact_match
    return next(
        (entry for entry in row_entries if _match_any_label(entry[1], candidates)),
        None,
    )


def extract_structured_financials(storage_path: str, file_type: str) -> Optional[dict[str, Any]]:
    """Deterministically extract key financial rows from spreadsheet-style source docs."""
    if file_type.lower() != "xlsx":
        return None

    path = Path(storage_path)
    if not path.exists():
        return None

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_name = max(wb.sheetnames, key=_sheet_priority)
        if _sheet_priority(sheet_name) <= 0:
            wb.close()
            return None

        ws = wb[sheet_name]
        fy_cols = _find_fy_columns(ws)
        if len(fy_cols) < 3:
            wb.close()
            return None

        header_snippets = []
        for row in ws.iter_rows(min_row=1, max_row=4, max_col=6, values_only=True):
            header_snippets.extend(str(cell or "") for cell in row if cell is not None)
        header_snippets.append(sheet_name)
        reporting_unit, multiplier = _detect_reporting_unit(header_snippets)

        row_entries: list[tuple[int, str, tuple[Any, ...]]] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            label = _normalize_label(row[0] if row else "")
            if label:
                row_entries.append((row_idx, label, row))

        revenue_row = _pick_row(row_entries, _SPREADSHEET_REVENUE_LABELS)
        operating_income_row = _pick_row(row_entries, _SPREADSHEET_OPERATING_INCOME_LABELS)
        depreciation_rows = [entry for entry in row_entries if _match_any_label(entry[1], _SPREADSHEET_DEPRECIATION_LABELS)]
        cash_rows = [entry for entry in row_entries if _match_any_label(entry[1], _SPREADSHEET_CASH_LABELS)]
        debt_rows = [entry for entry in row_entries if _match_any_label(entry[1], _SPREADSHEET_DEBT_LABELS)]
        pat_row = _pick_row(row_entries, _SPREADSHEET_PAT_LABELS)
        eps_row = _pick_row(row_entries, _SPREADSHEET_EPS_LABELS)
        shares_row = _pick_row(row_entries, _SPREADSHEET_SHARES_LABELS)

        if revenue_row is None:
            wb.close()
            return None

        fiscal_years = [year for _, year in fy_cols]
        fiscal_labels = [f"FY{str(year)[-2:]}" for year in fiscal_years]
        source_suffix = f"{sheet_name} | {fiscal_labels[0]}-{fiscal_labels[-1]} | unit={reporting_unit}"

        revenues = _series_from_row(revenue_row[2], fy_cols, multiplier)
        if not all(value is not None and value > 0 for value in revenues):
            wb.close()
            return None

        dep_total = _sum_series([(row_idx, row) for row_idx, _, row in depreciation_rows], fy_cols, multiplier)
        operating_income = (
            _series_from_row(operating_income_row[2], fy_cols, multiplier)
            if operating_income_row is not None
            else [None for _ in fy_cols]
        )
        ebitda_margins: list[Optional[float]] = []
        for revenue, op_income, dep in zip(revenues, operating_income, dep_total):
            if revenue and op_income is not None:
                ebitda_margins.append((op_income + dep) / revenue)
            else:
                ebitda_margins.append(None)

        cash_and_equivalents = sum(_sum_series([(row_idx, row) for row_idx, _, row in cash_rows], fy_cols, multiplier))
        cash_latest = _sum_series([(row_idx, row) for row_idx, _, row in cash_rows], fy_cols, multiplier)[-1]
        total_borrowings = _sum_series([(row_idx, row) for row_idx, _, row in debt_rows], fy_cols, multiplier)[-1]
        profit_after_tax = (
            _series_from_row(pat_row[2], fy_cols, multiplier)[-1]
            if pat_row is not None
            else None
        )
        basic_eps = (
            _series_from_row(eps_row[2], fy_cols, 1.0)[-1]
            if eps_row is not None
            else None
        )
        shares_outstanding = (
            _series_from_row(shares_row[2], fy_cols, 1.0)[-1]
            if shares_row is not None
            else None
        )

        industry_sector = None
        if any("information technology and consultancy services" in label for _, label, _ in row_entries):
            industry_sector = "IT Services"

        latest_revenue = revenues[-1]
        latest_dep = dep_total[-1] if dep_total else 0.0
        da_percent_rev = (latest_dep / latest_revenue) if latest_revenue else None
        net_debt = total_borrowings - cash_latest

        extracted_data = {
            "historical_revenues": [int(value) for value in revenues if value is not None],
            "historical_ebitda_margins": [round(float(value), 4) for value in ebitda_margins if value is not None],
            "net_debt": round(float(net_debt), 2),
            "total_borrowings": round(float(total_borrowings), 2),
            "cash_and_equivalents": round(float(cash_latest), 2),
            "shares_outstanding": round(float(shares_outstanding), 0) if shares_outstanding else None,
            "diluted_shares_outstanding": round(float(shares_outstanding), 0) if shares_outstanding else None,
            "cap_ex_percent_rev": None,
            "da_percent_rev": round(float(da_percent_rev), 4) if da_percent_rev is not None else None,
            "debt_to_equity": None,
            "beta": None,
            "base_fy": fiscal_years[-1],
            "reporting_unit": reporting_unit,
            "industry_sector": industry_sector,
            "profit_after_tax": round(float(profit_after_tax), 2) if profit_after_tax is not None else None,
            "basic_eps": round(float(basic_eps), 2) if basic_eps is not None else None,
            "currency": "INR",
            "listing_status": "listed" if shares_outstanding and basic_eps else "unknown",
        }

        audit_trail = [
            {
                "field": "historical_revenues",
                "value": extracted_data["historical_revenues"],
                "confidence": 0.95,
                "source_citation": f"{source_suffix} | row {revenue_row[0]} ({revenue_row[1]})",
                "reasoning": "Pulled directly from FY total columns in the INR financial sheet.",
            },
            {
                "field": "historical_ebitda_margins",
                "value": extracted_data["historical_ebitda_margins"],
                "confidence": 0.90,
                "source_citation": (
                    f"{source_suffix} | operating income row {operating_income_row[0] if operating_income_row else 'n/a'}"
                    f" + depreciation rows {[row_idx for row_idx, _, _ in depreciation_rows]}"
                ),
                "reasoning": "Computed as (Operating Income + summed Depreciation rows) / Total Revenue for each fiscal year.",
            },
            {
                "field": "cash_and_equivalents",
                "value": extracted_data["cash_and_equivalents"],
                "confidence": 0.90,
                "source_citation": f"{source_suffix} | cash-like rows {[row_idx for row_idx, _, _ in cash_rows]}",
                "reasoning": "Summed cash, investments, and bank deposit rows for the latest fiscal year.",
            },
            {
                "field": "total_borrowings",
                "value": extracted_data["total_borrowings"],
                "confidence": 0.85,
                "source_citation": f"{source_suffix} | borrowing rows {[row_idx for row_idx, _, _ in debt_rows]}",
                "reasoning": "Summed borrowing rows for the latest fiscal year.",
            },
        ]

        if extracted_data["shares_outstanding"] is not None:
            audit_trail.extend(
                [
                    {
                        "field": "shares_outstanding",
                        "value": extracted_data["shares_outstanding"],
                        "confidence": 0.92,
                        "source_citation": f"{source_suffix} | row {shares_row[0]} ({shares_row[1]})",
                        "reasoning": "Used weighted average diluted/basic shares from the spreadsheet.",
                    },
                    {
                        "field": "diluted_shares_outstanding",
                        "value": extracted_data["diluted_shares_outstanding"],
                        "confidence": 0.92,
                        "source_citation": f"{source_suffix} | row {shares_row[0]} ({shares_row[1]})",
                        "reasoning": "Workbook provides a single weighted average diluted/basic share count.",
                    },
                ]
            )
        if extracted_data["profit_after_tax"] is not None:
            audit_trail.append(
                {
                    "field": "profit_after_tax",
                    "value": extracted_data["profit_after_tax"],
                    "confidence": 0.90,
                    "source_citation": f"{source_suffix} | row {pat_row[0]} ({pat_row[1]})" if pat_row else source_suffix,
                    "reasoning": "Pulled from the net profit after tax row for the latest fiscal year.",
                }
            )
        if extracted_data["basic_eps"] is not None:
            audit_trail.append(
                {
                    "field": "basic_eps",
                    "value": extracted_data["basic_eps"],
                    "confidence": 0.90,
                    "source_citation": f"{source_suffix} | row {eps_row[0]} ({eps_row[1]})" if eps_row else source_suffix,
                    "reasoning": "Pulled from the EPS disclosure row for the latest fiscal year.",
                }
            )
        if extracted_data["da_percent_rev"] is not None:
            audit_trail.append(
                {
                    "field": "da_percent_rev",
                    "value": extracted_data["da_percent_rev"],
                    "confidence": 0.85,
                    "source_citation": f"{source_suffix} | depreciation rows {[row_idx for row_idx, _, _ in depreciation_rows]}",
                    "reasoning": "Calculated using summed depreciation rows divided by latest revenue.",
                }
            )
        if extracted_data["industry_sector"] is not None:
            audit_trail.append(
                {
                    "field": "industry_sector",
                    "value": extracted_data["industry_sector"],
                    "confidence": 0.85,
                    "source_citation": f"{sheet_name} | segment/service line labels",
                    "reasoning": "Detected explicit Information Technology and Consultancy Services labeling in the workbook.",
                }
            )

        wb.close()
        return {
            "extracted_data": extracted_data,
            "audit_trail": audit_trail,
            "reconciliation_log": "Structured spreadsheet extraction from FY columns in the INR financial sheet.",
            "extraction_mode": "structured_spreadsheet",
        }
    except Exception as exc:
        logger.warning("Structured spreadsheet extraction failed for %s: %s", path.name, exc)
        return None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _nodes_to_text(nodes) -> str:
    """Join LlamaIndex Document/Node objects into one string."""
    parts = []
    for node in nodes:
        text = getattr(node, "text", None) or getattr(node, "get_content", lambda: "")()
        if text and text.strip():
            parts.append(text.strip())
    return "\n\n".join(parts)


def _clean_text(text: str) -> str:
    """Normalize parser output while preserving line structure for tables."""
    if not text:
        return ""

    # Replace NULs and normalize common whitespace issues from OCR/PDFs.
    text = text.replace("\x00", " ").replace("\r\n", "\n").replace("\r", "\n")

    cleaned_lines = []
    for line in text.split("\n"):
        # Collapse internal whitespace but keep line breaks.
        collapsed = " ".join(line.split())
        if collapsed:
            cleaned_lines.append(collapsed)

    return "\n".join(cleaned_lines).strip()


def _clean_table_text(text: str) -> str:
    """Normalize table-heavy text while preserving tab-delimited structure."""
    if not text:
        return ""

    text = text.replace("\x00", " ").replace("\r\n", "\n").replace("\r", "\n")
    cleaned_lines = []
    for line in text.split("\n"):
        if "\t" in line:
            cols = [" ".join(col.split()) for col in line.split("\t")]
            if any(cols):
                cleaned_lines.append("\t".join(cols).rstrip("\t"))
        else:
            collapsed = " ".join(line.split())
            if collapsed:
                cleaned_lines.append(collapsed)
    return "\n".join(cleaned_lines).strip()


def _truncate_text(text: str, max_chars: int = _MAX_DOC_CHARS) -> str:
    if len(text) <= max_chars:
        return text
    omitted = len(text) - max_chars
    return f"{text[:max_chars]}\n\n[...truncated {omitted:,} chars for context window safety...]"


def _safe_cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # Keep financial precision sane without noisy long tails.
        as_int = int(value)
        return str(as_int) if value == as_int else f"{value:.6g}"
    return str(value).strip()


_RAPIDOCR_ENGINE = None


def _ocr_page_with_tesseract(page) -> str:
    """OCR one PDF page via pytesseract. Returns empty string on failure."""
    try:
        import fitz  # type: ignore
        import pytesseract
        from PIL import Image

        matrix = fitz.Matrix(2, 2)
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        img = Image.open(io.BytesIO(pix.tobytes("png")))

        ocr_lang = os.environ.get("AIBAA_OCR_LANG", "eng")
        ocr_config = os.environ.get("AIBAA_OCR_CONFIG", "--oem 3 --psm 6")
        text = pytesseract.image_to_string(img, lang=ocr_lang, config=ocr_config)
        return _clean_text(text)
    except Exception as exc:
        logger.warning("Tesseract OCR unavailable/failed: %s", exc)
        return ""


def _ocr_page_with_rapidocr(page) -> str:
    """OCR one PDF page via RapidOCR (pure Python/ONNX runtime)."""
    global _RAPIDOCR_ENGINE
    try:
        import fitz  # type: ignore
        import numpy as np
        from PIL import Image
        from rapidocr_onnxruntime import RapidOCR

        if _RAPIDOCR_ENGINE is None:
            _RAPIDOCR_ENGINE = RapidOCR()

        matrix = fitz.Matrix(2, 2)
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
        img_arr = np.array(img)

        result, _ = _RAPIDOCR_ENGINE(img_arr)
        if not result:
            return ""

        lines = []
        for item in result:
            # RapidOCR tuple shape: [box, text, score]
            if len(item) >= 2:
                text = str(item[1]).strip()
                if text:
                    lines.append(text)
        return _clean_text("\n".join(lines))
    except Exception as exc:
        logger.warning("RapidOCR fallback failed: %s", exc)
        return ""


def _ocr_page_with_fallback(page) -> str:
    """Try OCR backends in order: Tesseract -> RapidOCR."""
    text = _ocr_page_with_tesseract(page)
    if text:
        return text
    return _ocr_page_with_rapidocr(page)


def _post_process_financial_text(text: str) -> str:
    """Prepend a compact financial-signals block for key valuation fields."""
    if not text:
        return text

    section_hits: dict[str, list[str]] = {k: [] for k in _SECTION_PATTERNS}
    section_seen: dict[str, set[str]] = {k: set() for k in _SECTION_PATTERNS}

    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        if not _NUMERIC_LINE_RE.search(line):
            continue

        for section, patterns in _SECTION_PATTERNS.items():
            if any(p.search(line) for p in patterns):
                key = line.lower()
                if key not in section_seen[section]:
                    section_seen[section].add(key)
                    section_hits[section].append(line)
                break

    summary_lines = ["### Financial Signals (Revenue, EBITDA, Debt, Cash)"]
    any_signal = False
    for section in ("Revenue", "EBITDA", "Debt", "Cash"):
        hits = section_hits[section][:15]
        if not hits:
            continue
        any_signal = True
        summary_lines.append(f"#### {section}")
        for item in hits:
            summary_lines.append(f"- {item}")

    if not any_signal:
        return text

    summary = "\n".join(summary_lines)
    return f"{summary}\n\n{text}"


# ---------------------------------------------------------------------------
# Per-format parsers
# ---------------------------------------------------------------------------

def _parse_pdf(path: Path) -> str:
    # Use PyMuPDF directly — much faster and more reliable than pypdf for
    # complex financial PDFs.  LlamaIndex PDFReader wraps pypdf which hangs
    # on large / scanned PDFs, so we skip it here.
    try:
        import fitz  # type: ignore  (PyMuPDF)
        doc = fitz.open(str(path))
        pages = []
        ocr_pages = 0
        for idx, page in enumerate(doc, start=1):
            # Primary extraction.
            page_text = page.get_text("text") or ""

            # Fallback for pages where plain-text mode returns sparse output.
            if len(page_text.strip()) < 80:
                blocks = page.get_text("blocks") or []
                block_lines = []
                for block in blocks:
                    if len(block) >= 5:
                        block_text = str(block[4]).strip()
                        if block_text:
                            block_lines.append(block_text)
                page_text = "\n".join(block_lines)

            # OCR fallback for scanned/low-text pages.
            if len(page_text.strip()) < _OCR_MIN_PAGE_CHARS:
                ocr_text = _ocr_page_with_fallback(page)
                if ocr_text:
                    page_text = ocr_text
                    ocr_pages += 1

            cleaned = _clean_text(page_text)
            if cleaned:
                pages.append(f"## Page {idx}\n{cleaned}")

        # Free native resources ASAP.
        doc.close()
        text = "\n\n".join(pages)
        if text.strip():
            if ocr_pages:
                logger.info("PDF OCR fallback used for %d pages in %s", ocr_pages, path.name)
            return _truncate_text(_post_process_financial_text(text))
    except Exception as exc:
        logger.error("PyMuPDF PDF parse failed: %s", exc)

    return ""


def _parse_excel(path: Path) -> str:
    # openpyxl-first for better sheet/header fidelity in financial models.
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                cells = [_safe_cell_to_str(c) for c in row]
                if any(c for c in cells):
                    rows.append("\t".join(cells))
                    row_count += 1
                    if row_count >= _MAX_EXCEL_ROWS_PER_SHEET:
                        rows.append("[...rows truncated for context window safety...]")
                        break
            if rows:
                sheets.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))
        wb.close()
        text = "\n\n".join(sheets)
        if text.strip():
            return _truncate_text(_post_process_financial_text(_clean_table_text(text)))
    except Exception as exc:
        logger.warning("openpyxl Excel parse failed (%s), trying LlamaIndex", exc)

    # Fallback: LlamaIndex reader
    try:
        from llama_index.readers.file import PandasExcelReader
        reader = PandasExcelReader(concat_rows=False)
        docs = reader.load_data(file=path)
        text = _nodes_to_text(docs)
        text = _clean_table_text(text)
        if text.strip():
            return _truncate_text(_post_process_financial_text(text))
    except Exception as exc:
        logger.error("LlamaIndex PandasExcelReader failed: %s", exc)

    return ""


def _parse_xls(path: Path) -> str:
    """Legacy .xls via pandas (xlrd)."""
    try:
        import pandas as pd
        xl = pd.ExcelFile(str(path))
        parts = []
        for sheet in xl.sheet_names:
            df = xl.parse(sheet)
            if len(df) > _MAX_EXCEL_ROWS_PER_SHEET:
                df = df.head(_MAX_EXCEL_ROWS_PER_SHEET)
            parts.append(f"## Sheet: {sheet}\n{df.to_string(index=False)}")
        return _truncate_text(_post_process_financial_text(_clean_table_text("\n\n".join(parts))))
    except Exception as exc:
        logger.error(".xls parse failed: %s", exc)
        return ""


def _parse_csv(path: Path) -> str:
    try:
        from llama_index.readers.file import PandasCSVReader
        reader = PandasCSVReader()
        docs = reader.load_data(file=path)
        text = _nodes_to_text(docs)
        if text.strip():
            return _truncate_text(_post_process_financial_text(_clean_table_text(text)))
    except Exception as exc:
        logger.warning("LlamaIndex PandasCSVReader failed (%s), falling back to pandas", exc)

    try:
        import pandas as pd
        df = pd.read_csv(str(path))
        if len(df) > _MAX_EXCEL_ROWS_PER_SHEET:
            df = df.head(_MAX_EXCEL_ROWS_PER_SHEET)
        return _truncate_text(_post_process_financial_text(_clean_table_text(df.to_string(index=False))))
    except Exception as exc:
        logger.error("CSV parse failed: %s", exc)
        return ""


def _parse_docx(path: Path) -> str:
    try:
        from llama_index.readers.file import DocxReader
        reader = DocxReader()
        docs = reader.load_data(file=path)
        text = _nodes_to_text(docs)
        if text.strip():
            return text
    except Exception as exc:
        logger.warning("LlamaIndex DocxReader failed (%s), falling back to python-docx", exc)

    try:
        from docx import Document as DocxDocument  # type: ignore
        doc = DocxDocument(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return _truncate_text(_clean_text("\n\n".join(paragraphs)))
    except Exception as exc:
        logger.error("python-docx parse failed: %s", exc)
        return ""


def _parse_text(path: Path) -> str:
    try:
        return _truncate_text(_post_process_financial_text(_clean_text(path.read_text(encoding="utf-8", errors="replace"))))
    except Exception as exc:
        logger.error("Text read failed: %s", exc)
        return ""


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

_PARSERS = {
    "pdf":  _parse_pdf,
    "xlsx": _parse_excel,
    "xls":  _parse_xls,
    "csv":  _parse_csv,
    "docx": _parse_docx,
    "txt":  _parse_text,
    "json": _parse_text,
}


def parse_document(storage_path: str, file_type: str) -> Optional[str]:
    """
    Parse a document file and return its text content.

    Parameters
    ----------
    storage_path : absolute path to the file on disk
    file_type    : lowercase extension without dot (pdf, xlsx, csv, ...)

    Returns the extracted text, or None if parsing failed.
    """
    path = Path(storage_path)
    if not path.exists():
        logger.error("parse_document: file not found: %s", storage_path)
        return None

    parser = _PARSERS.get(file_type.lower())
    if parser is None:
        logger.warning("parse_document: unsupported type '%s'", file_type)
        return None

    try:
        text = parser(path)
        logger.info(
            "parse_document: parsed %s (%s) → %d chars",
            path.name, file_type, len(text)
        )
        return text if text else None
    except Exception as exc:
        logger.exception("parse_document: unexpected error for %s: %s", storage_path, exc)
        return None
