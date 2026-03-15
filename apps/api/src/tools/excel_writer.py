import logging
import os
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
_UNSAFE_CHARS = re.compile(r"[^a-z0-9_\-]")

# Canonical output directory: apps/data/outputs (4 levels up from this file)
_DEFAULT_OUTPUT_DIR = str(
    Path(__file__).resolve().parent.parent.parent.parent / "data" / "outputs"
)


class WorkbookBuilder:
    """Tool to generate formatting-compliant IB Excel Models."""

    def __init__(self, output_dir: str = _DEFAULT_OUTPUT_DIR):
        self.output_dir = os.path.abspath(output_dir)
        os.makedirs(self.output_dir, exist_ok=True)
        
        # IB Standard Formatting
        self.header_font = Font(name='Arial', bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        self.calc_font = Font(name='Arial', color='0000FF') # Blue for calculated
        self.hardcode_font = Font(name='Arial', color='000000') # Black for hardcoded
        self.hist_font = Font(name='Arial', color='008000')  # Green for historical actuals
        self.align_center = Alignment(horizontal='center')
        self.pct_font = Font(name='Arial', color='808080', italic=True)  # Grey for % rows
        
        thin_border = Side(border_style="thin", color="000000")
        self.border_bottom = Border(bottom=thin_border)
        self.border_top = Border(top=thin_border)
        
    def _apply_header_style(self, cell):
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.align_center
        
    def write_dcf_model(self, deal_name: str, assumptions: Dict[str, float], 
                        projections: Dict[str, List[float]], valuation: Dict[str, Any], 
                        currency: str = "INR", historical: Dict[str, Any] = None) -> str:
        """Generates a multi-tab DCF Workbook and saves it to disk."""
        wb = openpyxl.Workbook()
        is_private_company = bool(valuation.get("is_private_company"))
        valuation_basis = valuation.get("valuation_basis", "share_price")
        show_per_share_rows = (not is_private_company) and valuation_basis == "share_price"
        liquidity_discount = float(valuation.get("liquidity_discount", 0) or 0)
        control_premium = float(valuation.get("control_premium", 0) or 0)
        
        # Dynamic currency format
        currency_symbol = "₹" if currency == "INR" else "$"
        curr_fmt = f'{currency_symbol}#,##0.00'
        
        # FY labels from projections (e.g. FY2026E, FY2027E...)
        fy_labels = projections.get("fy_labels", [f"Year {i+1}" for i in range(5)])
        hist_labels = historical.get("fy_labels", []) if historical else []
        num_hist = len(hist_labels)
        num_proj = len(fy_labels)
        
        # --- Tab 1: Assumptions ---
        ws_assumptions = wb.active
        ws_assumptions.title = "Assumptions"
        ws_assumptions["A1"] = f"{deal_name} - Key Assumptions"
        ws_assumptions["A1"].font = Font(name='Arial', size=14, bold=True)
        
        r = 3
        for k, v in assumptions.items():
            if k == "base_fy":
                continue  # Don't show internal parameter
            if isinstance(v, dict):
                continue  # Skip nested dicts like working_capital_days
            ws_assumptions.cell(row=r, column=1, value=k.replace("_", " ").title())
            c = ws_assumptions.cell(row=r, column=2, value=v)
            c.font = self.hardcode_font
            c.number_format = '0.00%' if isinstance(v, (int, float)) and v < 1 else '#,##0.00'
            r += 1
            
        r += 1
        ws_assumptions.cell(row=r, column=1, value="WACC")
        ws_assumptions.cell(row=r, column=2, value=valuation["wacc"]).number_format = '0.00%'
        r += 1
        ws_assumptions.cell(row=r, column=1, value="Terminal Growth Rate")
        ws_assumptions.cell(row=r, column=2, value=valuation["terminal_growth_rate"]).number_format = '0.00%'
        
        # --- WACC Breakdown Sub-Section ---
        wacc_bk = valuation.get("wacc_breakdown", {})
        if wacc_bk:
            r += 2
            ws_assumptions.cell(row=r, column=1, value="WACC Calculation Breakdown").font = Font(name='Arial', bold=True, italic=True)
            if "note" in wacc_bk:
                r += 1
                ws_assumptions.cell(row=r, column=1, value=wacc_bk["note"]).font = Font(name='Arial', color='808080')
            else:
                if wacc_bk.get("method") == "build_up":
                    wacc_fields = [
                        ("Risk-Free Rate", "risk_free_rate", True),
                        ("Equity Risk Premium", "equity_risk_premium", True),
                        ("Size Premium", "size_premium", True),
                        ("Specific Risk Premium", "specific_risk_premium", True),
                        ("Cost of Equity (Build-Up)", "cost_of_equity", True),
                        ("Cost of Debt", "cost_of_debt", True),
                        ("After-Tax Cost of Debt", "after_tax_cost_of_debt", True),
                        ("Weight of Equity", "weight_of_equity", True),
                        ("Weight of Debt", "weight_of_debt", True)
                    ]
                else:
                    wacc_fields = [
                        ("Risk-Free Rate (10Y G-Sec)", "risk_free_rate", True),
                        ("Equity Risk Premium", "equity_risk_premium", True),
                        ("Beta", "beta", False),
                        ("Size Premium", "size_premium", True),
                        ("Specific Risk Premium", "specific_risk_premium", True),
                        ("Cost of Equity (CAPM)", "cost_of_equity", True),
                        ("Cost of Debt", "cost_of_debt", True),
                        ("After-Tax Cost of Debt", "after_tax_cost_of_debt", True),
                        ("Weight of Equity", "weight_of_equity", True),
                        ("Weight of Debt", "weight_of_debt", True)
                    ]
                for label, dict_key, is_pct in wacc_fields:
                    if dict_key in wacc_bk:
                        r += 1
                        ws_assumptions.cell(row=r, column=1, value=label)
                        c = ws_assumptions.cell(row=r, column=2, value=wacc_bk[dict_key])
                        c.font = self.calc_font
                        c.number_format = '0.00%' if is_pct else '0.00'
                        
        ws_assumptions.column_dimensions['A'].width = 40
        
        # --- Tab 2: Projections ---
        ws_proj = wb.create_sheet(title="Projections")
        ws_proj["A1"] = f"{deal_name} - Financial Projections ({currency})"
        ws_proj["A1"].font = Font(name='Arial', size=14, bold=True)
        
        # Build headers: Metric | FY2023A | FY2024A | FY2025A | FY2026E | FY2027E | ...
        all_headers = ["Metric"] + hist_labels + fy_labels
        for col, header in enumerate(all_headers, start=1):
            cell = ws_proj.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)
        
        # Data start column: after Metric column
        hist_col_start = 2  # Column B
        proj_col_start = hist_col_start + num_hist  # After historical columns
        
        # Revenue row with historical + projected
        current_row = 4
        
        # Helper to write a metric row (historical + projected)
        def write_metric_row(row, label, hist_vals, proj_vals, is_pct=False, is_ufcf=False):
            cell_label = ws_proj.cell(row=row, column=1, value=label)
            cell_label.font = Font(name='Arial', bold=True)
            if is_ufcf:
                cell_label.border = self.border_top
            
            # Historical values (green font)
            if hist_vals:
                for i, val in enumerate(hist_vals):
                    cell = ws_proj.cell(row=row, column=hist_col_start + i, value=val)
                    cell.font = self.hist_font
                    cell.number_format = '0.00%' if is_pct else '#,##0.00'
            
            # Projected values (blue font)
            for i, val in enumerate(proj_vals):
                cell = ws_proj.cell(row=row, column=proj_col_start + i, value=val)
                cell.font = self.pct_font if is_pct else self.calc_font
                cell.number_format = '0.00%' if is_pct else '#,##0.00'
                if is_ufcf:
                    cell.border = self.border_top
        
        # Historical EBITDA values
        hist_revenue = historical.get("revenue", []) if historical else []
        hist_ebitda = historical.get("ebitda", []) if historical else []
        hist_margins = historical.get("ebitda_margins", []) if historical else []
        
        # Compute historical revenue growth
        hist_growth = []
        for i in range(len(hist_revenue)):
            if i == 0:
                hist_growth.append(None)  # No prior year
            else:
                g = (hist_revenue[i] / hist_revenue[i-1] - 1) * 100 if hist_revenue[i-1] > 0 else 0
                hist_growth.append(round(g, 2))
        
        # Convert historical margins to percentages for display
        hist_margins_pct = [round(m * 100, 2) for m in hist_margins]
        
        # Write all metric rows
        write_metric_row(current_row, "Revenue", hist_revenue, projections.get("revenue", []))
        current_row += 1
        write_metric_row(current_row, "  Revenue Growth %", hist_growth, projections.get("revenue_growth_pct", []), is_pct=False)
        current_row += 1
        write_metric_row(current_row, "EBITDA", hist_ebitda, projections.get("ebitda", []))
        current_row += 1
        write_metric_row(current_row, "  EBITDA Margin %", hist_margins_pct, projections.get("ebitda_margin_pct", []), is_pct=False)
        current_row += 1
        write_metric_row(current_row, "Less: D&A", [], projections.get("da", []))
        current_row += 1
        write_metric_row(current_row, "EBIT", [], projections.get("ebit", []))
        current_row += 1
        write_metric_row(current_row, "Less: Taxes", [], projections.get("taxes", []))
        current_row += 1
        write_metric_row(current_row, "EBIAT", [], projections.get("ebiat", []))
        current_row += 1
        write_metric_row(current_row, "Plus: D&A", [], projections.get("da", []))
        current_row += 1
        write_metric_row(current_row, "Less: CapEx", [], projections.get("capex", []))
        current_row += 1
        write_metric_row(current_row, "Less: Change in NWC", [], projections.get("nwc_change", []))
        current_row += 1
        write_metric_row(current_row, "Unlevered Free Cash Flow", [], projections.get("ufcf", []), is_ufcf=True)
        
        ws_proj.column_dimensions['A'].width = 30
        for col_idx in range(2, 2 + num_hist + num_proj):
            ws_proj.column_dimensions[get_column_letter(col_idx)].width = 18
            
        # --- Tab 3: Valuation ---
        ws_val = wb.create_sheet(title="Valuation")
        ws_val["A1"] = f"{deal_name} - DCF Valuation ({currency})"
        ws_val["A1"].font = Font(name='Arial', size=14, bold=True)
        
        # Use FY labels for projected columns only
        val_headers = ["Metric"] + fy_labels
        for col, header in enumerate(val_headers, start=1):
            cell = ws_val.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)
            
        current_row = 4
        metrics_val = [
            ("Unlevered Free Cash Flow", projections.get("ufcf", [])),
            ("Discount Factor", valuation.get("discount_factors", [])),
            ("PV of Free Cash Flow", valuation.get("pv_of_fcf_array", []))
        ]
        
        for metric_name, values in metrics_val:
            ws_val.cell(row=current_row, column=1, value=metric_name).font = Font(name='Arial', bold=True)
            for col, val in enumerate(values, start=2):
                cell = ws_val.cell(row=current_row, column=col, value=val)
                cell.font = self.calc_font
                cell.number_format = '0.0000' if "Discount" in metric_name else '#,##0.00'
            current_row += 1
            
        current_row += 2
        
        # Equity Bridge with separate Borrowings and Cash
        net_debt_val = valuation.get("net_debt", 0)
        total_borrowings = valuation.get("total_borrowings")
        lease_liabilities = valuation.get("lease_liabilities", 0) or 0
        ccps_liability = valuation.get("ccps_liability", 0) or 0

        if total_borrowings is None:
            total_borrowings = abs(net_debt_val) * 1.2 if net_debt_val > 0 else 0

        total_debt = (total_borrowings or 0) + lease_liabilities + ccps_liability
        cash_and_equiv = total_debt - net_debt_val if net_debt_val > 0 else abs(net_debt_val) + total_debt
        
        equity_bridge = [
            ("Sum of PV of FCFs", valuation.get("pv_of_fcf_sum", 0)),
            ("PV of Terminal Value", valuation.get("pv_of_tv", 0)),
            ("Implied Enterprise Value", valuation.get("implied_enterprise_value", 0)),
            ("Less: Total Borrowings", total_borrowings),
            ("Less: Lease Liabilities", lease_liabilities),
            ("Less: CCPS Liability", ccps_liability),
            ("Add: Cash & Equivalents", cash_and_equiv),
            (f"Net Debt" if net_debt_val >= 0 else "Net Cash", abs(net_debt_val)),
        ]
        if is_private_company:
            equity_bridge.extend([
                ("Pre-Adjustment Equity Value", valuation.get("pre_private_adjustment_equity_value", valuation.get("implied_equity_value", 0))),
                ("Liquidity Discount %", liquidity_discount),
                ("Control Premium %", control_premium),
                ("Implied Equity Value", valuation.get("implied_equity_value", 0)),
            ])
        else:
            equity_bridge.append(("Implied Equity Value", valuation.get("implied_equity_value", 0)))
            if show_per_share_rows:
                equity_bridge.extend([
                    ("Shares Outstanding", valuation.get("shares_outstanding")),
                    ("Implied Share Price", valuation.get("implied_share_price")),
                ])

                # Add market price context if implied and market are available
                market_cap = valuation.get("market_cap")
                if market_cap and valuation.get("shares_outstanding"):
                    market_price = market_cap / valuation["shares_outstanding"]
                    equity_bridge.append(("Market Price (Reference)", round(market_price, 2)))
            else:
                equity_bridge.extend([
                    ("Shares Outstanding", valuation.get("shares_outstanding") if valuation.get("shares_outstanding") is not None else "Not verified"),
                    ("Implied Share Price", "Suppressed - share count unverified"),
                ])
        
        for name, val in equity_bridge:
            c_name = ws_val.cell(row=current_row, column=1, value=name)
            c_name.font = Font(name='Arial', bold=True)
            if "Enterprise" in name or "Equity Value" in name or "Share Price" in name:
                c_name.border = self.border_bottom
            c_val = ws_val.cell(row=current_row, column=2, value=val)
            c_val.font = self.calc_font
            if isinstance(val, str):
                c_val.number_format = '@'
            elif "Discount %" in name or "Premium %" in name:
                c_val.number_format = '0.00%'
            elif "Shares" in name:
                c_val.number_format = '#,##0'
            else:
                c_val.number_format = curr_fmt if "Price" in name or "Value" in name or "Sum" in name or "Borrowings" in name or "Cash" in name or "Net" in name or "Terminal" in name else '#,##0'
            if "Enterprise" in name or "Equity Value" in name or "Share Price" in name:
                c_val.border = self.border_bottom
            current_row += 1
            
        ws_val.column_dimensions['A'].width = 30
        for col_idx in range(2, 2 + num_proj):
            ws_val.column_dimensions[get_column_letter(col_idx)].width = 18

        # --- Tab 4: Sensitivity Analysis ---
        ws_sens = wb.create_sheet(title="Sensitivity Analysis")
        metric_title = "Implied Equity Value" if is_private_company or valuation_basis == "equity_value" else "Implied Share Price"
        ws_sens["A1"] = f"{deal_name} - {metric_title} Sensitivity ({currency})"
        ws_sens["A1"].font = Font(name='Arial', size=14, bold=True)
        
        ws_sens["A3"] = "Sensitivity: WACC vs. Terminal Growth Rate"
        ws_sens["A3"].font = Font(name="Arial", bold=True)
        
        metric_row = current_row - 1
        ws_sens["B4"] = f"='Valuation'!B{metric_row}"
        ws_sens["B4"].font = self.calc_font
        ws_sens["B4"].number_format = curr_fmt
        
        tgr_variations = [-0.01, -0.005, 0, 0.005, 0.01]
        wacc_variations = [-0.02, -0.01, 0, 0.01, 0.02]
        
        base_tgr = valuation["terminal_growth_rate"]
        base_wacc = valuation["wacc"]
        
        # Headers TGR
        for i, var in enumerate(tgr_variations):
            c = ws_sens.cell(row=4, column=3+i, value=base_tgr + var)
            c.number_format = "0.0%"
            c.font = Font(bold=True)
            
        # Headers WACC
        for i, var in enumerate(wacc_variations):
            c = ws_sens.cell(row=5+i, column=2, value=base_wacc + var)
            c.number_format = "0.0%"
            c.font = Font(bold=True)
            
        # Static sensitivity point calculation
        for r_idx, w_var in enumerate(wacc_variations):
            for c_idx, t_var in enumerate(tgr_variations):
                w_test = base_wacc + w_var
                t_test = base_tgr + t_var

                if w_test <= t_test:
                    metric_value = None
                    c = ws_sens.cell(row=5+r_idx, column=3+c_idx, value=metric_value)
                    c.font = self.calc_font
                    continue
                
                fcf_pv_sum = 0
                projs = projections.get("ufcf", [])
                for i, cf in enumerate(projs):
                    fcf_pv_sum += cf / ((1 + w_test) ** (i + 1))
                    
                tv = (projs[-1] * (1 + t_test)) / (w_test - t_test)
                tv_pv = tv / ((1 + w_test) ** len(projs))
                ev = fcf_pv_sum + tv_pv
                eq = ev - valuation.get("net_debt", 0)
                if is_private_company or valuation_basis == "equity_value":
                    metric_value = eq * (1 - liquidity_discount) * (1 + control_premium)
                else:
                    shares = valuation.get("shares_outstanding")
                    metric_value = eq / shares if shares and shares > 0 else None
                
                c = ws_sens.cell(row=5+r_idx, column=3+c_idx, value=metric_value)
                if metric_value is not None:
                    c.number_format = curr_fmt
                c.font = self.calc_font

        ws_sens.column_dimensions['A'].width = 10
        ws_sens.column_dimensions['B'].width = 15
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws_sens.column_dimensions[col].width = 15
            
        # Save File
        safe_name = _UNSAFE_CHARS.sub("_", deal_name.strip().lower().replace(" ", "_"))[:60]
        filename = f"dcf_model_{safe_name}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        # Guard: ensure path stays within output_dir
        if not os.path.abspath(filepath).startswith(self.output_dir):
            raise ValueError(f"Output path escapes output_dir: {filepath}")
        logger.info("[ExcelWriter] Saving DCF workbook to %s", filepath)

        wb.save(filepath)
        return filepath

    # ------------------------------------------------------------------
    # DD Checklist Workbook
    # ------------------------------------------------------------------

    def write_dd_checklist(self, deal_name: str, risk_data: dict) -> str:
        """
        Generate an IB-quality Due Diligence checklist Excel workbook.

        Tabs:
          1. Summary — overall risk score + red flags
          2. Risk Checklist — all risks by category, severity colour-coded
        """
        wb = openpyxl.Workbook()

        # ---- Colours ----
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        amber_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        white_font = Font(name="Arial", bold=True, color="FFFFFF")
        bold_font = Font(name="Arial", bold=True)
        normal_font = Font(name="Arial")
        wrap_align = Alignment(wrap_text=True, vertical="top")

        severity_fill = {
            "high": red_fill,
            "medium": amber_fill,
            "low": green_fill,
        }
        severity_font = {
            "high": white_font,
            "medium": Font(name="Arial", bold=True, color="FFFFFF"),
            "low": white_font,
        }

        # ---- TAB 1: Summary ----
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header
        ws_summary["A1"] = f"DUE DILIGENCE REPORT — {deal_name.upper()}"
        ws_summary["A1"].font = Font(name="Arial", bold=True, size=14)
        ws_summary["A1"].fill = self.header_fill
        ws_summary["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws_summary.merge_cells("A1:D1")

        ws_summary["A3"] = "Overall Risk Score"
        ws_summary["A3"].font = bold_font
        ws_summary["B3"] = risk_data.get("overall_risk_score", 0)
        ws_summary["C3"] = "/ 10"
        rating = risk_data.get("risk_rating", "MEDIUM")
        ws_summary["D3"] = rating
        rating_fill = {"LOW": green_fill, "MEDIUM": amber_fill, "HIGH": red_fill, "CRITICAL": red_fill}.get(rating, amber_fill)
        ws_summary["D3"].fill = rating_fill
        ws_summary["D3"].font = white_font

        ws_summary["A5"] = "Executive Summary"
        ws_summary["A5"].font = bold_font
        ws_summary["A6"] = risk_data.get("summary", "")
        ws_summary["A6"].alignment = wrap_align
        ws_summary.merge_cells("A6:D6")
        ws_summary.row_dimensions[6].height = 60

        ws_summary["A8"] = "RED FLAGS"
        ws_summary["A8"].font = Font(name="Arial", bold=True, color="FFFFFF")
        ws_summary["A8"].fill = red_fill
        ws_summary.merge_cells("A8:D8")

        headers = ["Flag", "Impact", "Recommendation"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_summary.cell(row=9, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill

        for row_idx, flag_item in enumerate(risk_data.get("red_flags", []), start=10):
            ws_summary.cell(row=row_idx, column=1, value=flag_item.get("flag", "")).font = normal_font
            ws_summary.cell(row=row_idx, column=2, value=flag_item.get("impact", "")).font = normal_font
            ws_summary.cell(row=row_idx, column=3, value=flag_item.get("recommendation", "")).font = normal_font

        for col in ["A", "B", "C", "D"]:
            ws_summary.column_dimensions[col].width = 35

        # ---- TAB 2: Risk Checklist ----
        ws_risks = wb.create_sheet("Risk Checklist")

        risk_headers = ["Category", "Risk Description", "Severity", "Evidence", "Mitigation", "Status", "Assigned To"]
        for col_idx, h in enumerate(risk_headers, start=1):
            cell = ws_risks.cell(row=1, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill

        row = 2
        categories = [
            ("Financial", risk_data.get("financial_risks", [])),
            ("Operational", risk_data.get("operational_risks", [])),
            ("Legal", risk_data.get("legal_risks", [])),
            ("Market", risk_data.get("market_risks", [])),
        ]
        for cat_name, risks in categories:
            for risk_item in risks:
                severity = str(risk_item.get("severity", "medium")).lower()
                ws_risks.cell(row=row, column=1, value=cat_name).font = normal_font
                ws_risks.cell(row=row, column=2, value=risk_item.get("risk", "")).alignment = wrap_align
                sev_cell = ws_risks.cell(row=row, column=3, value=severity.upper())
                sev_cell.fill = severity_fill.get(severity, grey_fill)
                sev_cell.font = severity_font.get(severity, bold_font)
                ws_risks.cell(row=row, column=4, value=risk_item.get("evidence", "")).alignment = wrap_align
                ws_risks.cell(row=row, column=5, value=risk_item.get("mitigation", "")).alignment = wrap_align
                ws_risks.cell(row=row, column=6, value="Open")
                ws_risks.cell(row=row, column=7, value="")
                row += 1

        col_widths = [15, 40, 10, 35, 35, 12, 20]
        for col_idx, width in enumerate(col_widths, start=1):
            ws_risks.column_dimensions[get_column_letter(col_idx)].width = width

        # Save
        safe_name = _UNSAFE_CHARS.sub("_", deal_name.strip().lower().replace(" ", "_"))[:60]
        filename = f"dd_checklist_{safe_name}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        if not os.path.abspath(filepath).startswith(self.output_dir):
            raise ValueError(f"Output path escapes output_dir: {filepath}")
        logger.info("[ExcelWriter] Saving DD checklist to %s", filepath)
        wb.save(filepath)
        return filepath

    # ------------------------------------------------------------------
    # LBO Model Workbook
    # ------------------------------------------------------------------

    def write_lbo_model(self, deal_name: str, lbo_result: dict) -> str:
        """
        Generate an IB-quality LBO model Excel workbook.

        Tabs:
          1. Sources & Uses
          2. Operating Model
          3. Debt Schedule
          4. Returns Analysis
          5. IRR Sensitivity
        """
        wb = openpyxl.Workbook()

        # ---- Shared fills ----
        green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        amber_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        light_blue_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        white_font = Font(name="Arial", bold=True, color="FFFFFF")
        bold_font = Font(name="Arial", bold=True)
        normal_font = Font(name="Arial")
        pct_fmt = "0.0%"
        num_fmt = '#,##0'
        curr_fmt = '#,##0.00'

        def hdr(ws, row, col, value, fill=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = self.header_font
            c.fill = fill or self.header_fill
            return c

        def val(ws, row, col, value, fmt=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = self.calc_font
            if fmt:
                c.number_format = fmt
            return c

        # ---- TAB 1: Sources & Uses ----
        ws_su = wb.active
        ws_su.title = "Sources & Uses"
        su = lbo_result.get("sources_uses", {})

        ws_su["A1"] = f"LBO MODEL — {deal_name.upper()} — SOURCES & USES"
        ws_su["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws_su["A1"].fill = self.header_fill
        ws_su.merge_cells("A1:C1")

        su_rows = [
            ("SOURCES", ""),
            ("Equity", su.get("equity", 0)),
            ("Senior Debt (TLA + TLB)", su.get("senior_debt", 0)),
            ("Mezzanine Debt", su.get("mezz_debt", 0)),
            ("Total Sources", su.get("total_sources", 0)),
            ("", ""),
            ("USES", ""),
            ("Purchase Price (Entry EV)", su.get("entry_ev", 0)),
            ("Total Uses", su.get("total_uses", 0)),
            ("", ""),
            ("Leverage Statistics", ""),
            ("Equity %", f"{su.get('equity_pct', 0):.1f}%"),
            ("Debt %", f"{su.get('debt_pct', 0):.1f}%"),
            ("Total Leverage (x EBITDA)", su.get("leverage_multiple", 0)),
        ]
        for r_idx, (label, value) in enumerate(su_rows, start=3):
            label_cell = ws_su.cell(row=r_idx, column=1, value=label)
            val_cell = ws_su.cell(row=r_idx, column=2, value=value)
            if label in ("SOURCES", "USES", "Leverage Statistics"):
                label_cell.font = bold_font
                label_cell.fill = light_blue_fill
            elif label in ("Total Sources", "Total Uses"):
                label_cell.font = bold_font
                val_cell.font = Font(name="Arial", bold=True, color="0000FF")
                val_cell.number_format = curr_fmt
            elif isinstance(value, (int, float)):
                val_cell.font = self.calc_font
                val_cell.number_format = curr_fmt

        ws_su.column_dimensions["A"].width = 30
        ws_su.column_dimensions["B"].width = 20

        # ---- TAB 2: Operating Model ----
        ws_ops = wb.create_sheet("Operating Model")
        op = lbo_result.get("operating_model", {})
        years = list(range(1, len(op.get("revenues", [])) + 1))

        ws_ops["A1"] = f"LBO MODEL — {deal_name.upper()} — OPERATING MODEL"
        ws_ops["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws_ops["A1"].fill = self.header_fill
        ws_ops.merge_cells(f"A1:{get_column_letter(len(years) + 1)}1")

        hdr(ws_ops, 2, 1, "Metric")
        for j, yr in enumerate(years, start=2):
            hdr(ws_ops, 2, j, f"Year {yr}")

        op_rows = [
            ("Revenue", op.get("revenues", []), curr_fmt),
            ("EBITDA", op.get("ebitda", []), curr_fmt),
            ("EBIT", op.get("ebit", []), curr_fmt),
            ("Unlevered FCF", op.get("ufcf", []), curr_fmt),
        ]
        for r_offset, (label, values, fmt) in enumerate(op_rows, start=3):
            ws_ops.cell(row=r_offset, column=1, value=label).font = bold_font
            for j, v in enumerate(values, start=2):
                val(ws_ops, r_offset, j, v, fmt)

        for col in range(1, len(years) + 2):
            ws_ops.column_dimensions[get_column_letter(col)].width = 18

        # ---- TAB 3: Debt Schedule ----
        ws_debt = wb.create_sheet("Debt Schedule")
        schedule = lbo_result.get("debt_schedule", [])

        ws_debt["A1"] = f"LBO MODEL — {deal_name.upper()} — DEBT SCHEDULE"
        ws_debt["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws_debt["A1"].fill = self.header_fill
        ws_debt.merge_cells(f"A1:I1")

        ds_headers = ["Year", "Opening Debt", "TLA Balance", "TLB Balance", "Mezz Balance",
                      "Cash Interest", "Mandatory Amort", "Cash Sweep", "Closing Debt"]
        for col_idx, h in enumerate(ds_headers, start=1):
            hdr(ws_debt, 2, col_idx, h)

        dscr_by_year = lbo_result.get("dscr_by_year", {})
        for r_offset, ds in enumerate(schedule, start=3):
            yr = ds.get("year", r_offset - 2)
            cells = [yr, ds.get("opening_debt"), ds.get("tla_balance"), ds.get("tlb_balance"),
                     ds.get("mezz_balance"), ds.get("cash_interest"), ds.get("mandatory_amort"),
                     ds.get("cash_sweep"), ds.get("closing_debt")]
            for col_idx, v in enumerate(cells, start=1):
                c = val(ws_debt, r_offset, col_idx, v, curr_fmt if col_idx > 1 else None)
            # DSCR annotation
            dscr = dscr_by_year.get(yr)
            if dscr is not None:
                dscr_cell = ws_debt.cell(row=r_offset, column=10, value=dscr)
                dscr_cell.number_format = "0.00x"
                if dscr >= 1.4:
                    dscr_cell.fill = green_fill
                    dscr_cell.font = Font(name="Arial", color="FFFFFF")
                elif dscr >= 1.0:
                    dscr_cell.fill = amber_fill
                    dscr_cell.font = Font(name="Arial", bold=True)
                else:
                    dscr_cell.fill = red_fill
                    dscr_cell.font = Font(name="Arial", bold=True, color="FFFFFF")

        hdr(ws_debt, 2, 10, "DSCR")
        for col in range(1, 11):
            ws_debt.column_dimensions[get_column_letter(col)].width = 18

        # ---- TAB 4: Returns Analysis ----
        ws_ret = wb.create_sheet("Returns Analysis")

        ws_ret["A1"] = f"LBO MODEL — {deal_name.upper()} — RETURNS ANALYSIS"
        ws_ret["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws_ret["A1"].fill = self.header_fill
        ws_ret.merge_cells("A1:C1")

        ret_rows = [
            ("Entry EV", lbo_result.get("entry_ev", 0), curr_fmt),
            ("Entry Equity", lbo_result.get("entry_equity", 0), curr_fmt),
            ("Total Debt at Entry", lbo_result.get("total_debt_at_entry", 0), curr_fmt),
            ("", None, None),
            ("Exit EV", lbo_result.get("exit_ev", 0), curr_fmt),
            ("Total Debt at Exit", lbo_result.get("total_debt_at_exit", 0), curr_fmt),
            ("Exit Equity", lbo_result.get("exit_equity", 0), curr_fmt),
            ("", None, None),
            ("IRR", lbo_result.get("irr", 0), pct_fmt),
            ("MOIC", lbo_result.get("moic", 0), "0.00x"),
        ]
        for r_idx, (label, value, fmt) in enumerate(ret_rows, start=3):
            ws_ret.cell(row=r_idx, column=1, value=label).font = bold_font
            if value is not None:
                c = ws_ret.cell(row=r_idx, column=2, value=value)
                c.font = self.calc_font
                if fmt:
                    c.number_format = fmt
            if label in ("IRR", "MOIC"):
                ws_ret.cell(row=r_idx, column=1).fill = blue_fill
                ws_ret.cell(row=r_idx, column=1).font = Font(name="Arial", bold=True, color="FFFFFF")

        ws_ret.column_dimensions["A"].width = 28
        ws_ret.column_dimensions["B"].width = 20

        # ---- TAB 5: IRR Sensitivity ----
        ws_sens = wb.create_sheet("Sensitivity")

        entry_multiples = [
            lbo_result["sources_uses"].get("leverage_multiple", 8.0) - 2,
            lbo_result["sources_uses"].get("leverage_multiple", 8.0) - 1,
            lbo_result["sources_uses"].get("leverage_multiple", 8.0),
            lbo_result["sources_uses"].get("leverage_multiple", 8.0) + 1,
            lbo_result["sources_uses"].get("leverage_multiple", 8.0) + 2,
        ]
        # Use entry EV/EBITDA for sensitivity (from sources_uses entry_ev / entry_ebitda)
        base_entry_ev = lbo_result.get("entry_ev", 0)
        # Approximate entry EV/EBITDA range using IRR result
        exit_base = lbo_result.get("exit_ev", base_entry_ev)
        base_irr = lbo_result.get("irr_pct", 20.0)

        ws_sens["A1"] = "IRR SENSITIVITY (Entry EV/EBITDA × Exit EV/EBITDA)"
        ws_sens["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws_sens["A1"].fill = self.header_fill
        ws_sens.merge_cells("A1:F1")

        ws_sens["A2"] = "IRR %"
        ws_sens["A2"].font = bold_font

        # Populate if sensitivity data is available
        sensitivity = lbo_result.get("irr_sensitivity", {})
        if sensitivity:
            exit_multiples = sorted({xm for row_dict in sensitivity.values() for xm in row_dict.keys()})
            hdr(ws_sens, 3, 1, "Entry \\ Exit")
            for col_idx, xm in enumerate(exit_multiples, start=2):
                hdr(ws_sens, 3, col_idx, f"{xm}x")
            for r_idx, (em, row_dict) in enumerate(sensitivity.items(), start=4):
                ws_sens.cell(row=r_idx, column=1, value=f"{em}x").font = bold_font
                for col_idx, xm in enumerate(exit_multiples, start=2):
                    irr_val = row_dict.get(xm)
                    c = ws_sens.cell(row=r_idx, column=col_idx, value=irr_val)
                    if irr_val is not None:
                        c.number_format = "0.0%"
                        c.value = irr_val / 100  # convert to decimal for pct format
                        if irr_val >= 20:
                            c.fill = green_fill
                            c.font = Font(name="Arial", color="FFFFFF")
                        elif irr_val >= 10:
                            c.fill = amber_fill
                        else:
                            c.fill = red_fill
                            c.font = Font(name="Arial", color="FFFFFF")
        else:
            ws_sens["A3"] = "Sensitivity data not available — run LBO engine with irr_sensitivity() to populate."

        for col in range(1, 7):
            ws_sens.column_dimensions[get_column_letter(col)].width = 16

        # Save
        safe_name = _UNSAFE_CHARS.sub("_", deal_name.strip().lower().replace(" ", "_"))[:60]
        filename = f"lbo_model_{safe_name}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        if not os.path.abspath(filepath).startswith(self.output_dir):
            raise ValueError(f"Output path escapes output_dir: {filepath}")
        logger.info("[ExcelWriter] Saving LBO workbook to %s", filepath)
        wb.save(filepath)
        return filepath
