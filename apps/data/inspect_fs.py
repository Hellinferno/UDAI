import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\Lenovo\Downloads\AI Investment Banking Analyst Agent (AIBAA)\aibaa\apps\data\uploads\ddd68588-1d53-47ba-914e-b6a3a3ebb54a\7865bf0c-0650-475e-8153-0f3ef69feaa0_financial_statements_(1).xlsx')
print('Sheets:', wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n=== {sn} (rows={ws.max_row}, cols={ws.max_column}) ===')
    for r in range(1, min(ws.max_row+1, 60)):
        row_vals = []
        for c in range(1, min(ws.max_column+1, 8)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_vals.append(f'C{c}={repr(v)}')
        if row_vals:
            print(f'  R{r}: {" | ".join(row_vals)}')
